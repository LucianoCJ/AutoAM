import pandas as pd 
from openpyxl import Workbook 
from openpyxl.styles import Font, Alignment 
from openpyxl.styles import PatternFill
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils import range_boundaries
from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet.page import PageMargins
import re
import datetime
from xlsx2html import xlsx2html
from weasyprint import HTML
from bs4 import BeautifulSoup
import requests

##############################################################

# Obtener la fecha del próximo domingo en curso
today = datetime.date.today()
#Domingo proximo
next_sunday = today + datetime.timedelta(days=(6-today.weekday()+7)%7)
#domingo anterior
before_sunday = today - datetime.timedelta(days=today.weekday() + 1)
#before_sunday = today - datetime.timedelta(days=(6-today.weekday()-7)%7)

# Dar formato a la fecha como mes-día-año
formatted_date = today.strftime("%b %d, %Y").replace("{0:0>2}".format(today.day), str(today.day) + ("th" if 11<= today.day <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(today.day % 10, 'th')),1)
formatted_date2 = next_sunday.strftime("%b %d").replace("{0:0>2}".format(next_sunday.day), str(next_sunday.day) + ("th" if 11<= next_sunday.day <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(next_sunday.day % 10, 'th')),1)
formatted_date3 = before_sunday.strftime("%b %d").replace("{0:0>2}".format(before_sunday.day), str(before_sunday.day) + ("th" if 11<= before_sunday.day <= 13 else {1: 'st', 2: 'nd', 3: 'rd'}.get(before_sunday.day % 10, 'th')),1)

#Fecha del archivo
formatted_date4 = next_sunday.strftime("%b%d%Y")

############################################################
originFile = 'ReporteOAGLATAM_20230312.csv'
# Leer el archivo CSV
df = pd.read_csv(originFile, header = None)

# Crear un nuevo libro de trabajo de Excel
wb = Workbook()

# Seleccionar la hoja activa
ws = wb.active

# Establecer el ancho de las columnas
ws.column_dimensions['A'].width = 3.9
ws.column_dimensions['B'].width = 6.75 
ws.column_dimensions['C'].width = 3.61     #COLUMNA 1
ws.column_dimensions['D'].width = 7.32     #COLUMNA 2
ws.column_dimensions['E'].width = 7.04     #COLUMNA 3
ws.column_dimensions['F'].width = 4.32 
ws.column_dimensions['G'].width = 4.32    #COLUMNA 4
ws.column_dimensions['H'].width = 4.04    #COLUMNA 5
ws.column_dimensions['I'].width = 3.89    #COLUMNA 6
ws.column_dimensions['J'].width = 4.32     #COLUMNA 7
ws.column_dimensions['K'].width = 4.18     #COLUMNA 8
ws.column_dimensions['L'].width = 5.32     #COLUMNA 9
ws.column_dimensions['M'].width = 6.04     #COLUMNA 10

# Definiendo los nuevos margenes modificar los margenes en pulgadas a cm.
nuevos_margenes = PageMargins(left = 0.23622047244094, right = 0.23622047244094, top = 0.86614173228346, bottom = 0.74803149606299, header = 0.31496062992126, footer = 0.31496062992126)

# Actualizando los margenes de la hoja de trabajo
ws.page_margins = nuevos_margenes

# Centrar los márgenes horizontal y verticalmente y activa las opciones de impresión
ws.print_options.horizontalCentered = True
ws.print_options.verticalCentered = True

# Establecer la altura deseada en la fila y columna especificada
ws.row_dimensions[1].height = 17.25

# Establecer el estilo de fuente y alineación
header_font = Font(name='Calibri', size=6, bold=True)
header_alignment = Alignment(horizontal='center', vertical='center')
cell_font = Font(name='Calibri', size=6)
cell_alignment = Alignment(horizontal='left', vertical='center')

# Crear una lista con los nombres de las cabeceras
cabeceras = ['CA', 'Market', 'Ind AM', 'Region', 'Month', 'Chg', 'Prev Ops', 'New Ops', 'Ops Chg', 'Prev Seat', 'New Seat', 'Seat Chg', '%_Seat Chg']

headers = cabeceras                             #Indico el número de columna a iniciar las cabeceras.
for col_num, header_title in enumerate(headers, 3):
    cell = ws.cell(row=1, column=col_num, value=header_title)
    cell.font = header_font
    cell.alignment = header_alignment

# Justifica el texto de los encabezados
for idx, cell in enumerate(ws[1],1):
    if idx<9:  
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    else:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Escribir los datos                
for row_num, row_data in enumerate(df.values, 2):
    for col_num, cell_value in enumerate(row_data, 2):
        cell = ws.cell(row=row_num, column=col_num, value=cell_value)
        cell.font = cell_font
        if col_num < 9:
            cell.alignment = cell_alignment
        else:
            cell.alignment = Alignment(horizontal='right', vertical='center')
    if str(ws.cell(row=row_num, column=4).value) == 'nan':
        ws.delete_rows(row_num)
        
#Definir altura de las celdas
for idx, value in enumerate(ws.iter_rows(),2):
    ws.row_dimensions[idx].height = 7.5
for idx, value in enumerate(ws.iter_rows(),ws.max_row):
    ws.row_dimensions[idx].height = 7.5
    if idx >= ws.max_row+int(ws.max_row/79)+1:
        break

# definir los colores para los renglones
color1 = 'F2F2F2'
color2 = 'FFFFFF'
change = False

# cambiar el formato de color de los renglones 
for idx, row in enumerate(ws.iter_rows(),1):
    if ((str(ws.cell(row=idx, column=3).value) == 'None') or (str(ws.cell(row=idx, column=4).value) == 'TOTAL')):
        fill = PatternFill(start_color=color2, end_color=color2, fill_type='solid')
        change = True              
    elif change:
        fill = PatternFill(start_color=color1, end_color=color1, fill_type='solid')
        change = False                    
    else:
        fill = PatternFill(start_color=color2, end_color=color2, fill_type='solid')
        change = True
    for cell in row:
        cell.fill = fill
        
# Poner el renglon de total en negritas y porcentajes negativos en rojo
for idx, row in enumerate(ws.iter_rows(),1):
    if str(ws.cell(row=idx, column=4).value) == 'TOTAL':
        for cell in row:
            cell.font = Font(bold=True, name='Calibri', size=6)
        if re.search("-", str(ws.cell(row=idx, column=15).value)):
            ws.cell(row=idx, column=15).font = Font(color = "FF0000", name='Calibri', size=6, bold=True)
    elif re.search("-", str(ws.cell(row=idx, column=15).value)):
         ws.cell(row=idx, column=15).font = Font(color = "FF0000", name='Calibri', size=6)

# Establecer color de relleno para los encabezados
fill = PatternFill(start_color='BFBFBF', end_color='BFBFBF', fill_type='solid')

# Justifica el texto de los encabezados
for cell in ws[1]:
    cell.fill = fill    #Agrego el color de relleno en el renglon 1

#Convirtiendo a Decimales
# Selecciona la columna que deseas convertir (por ejemplo, columna A)
columna = ws['O']

# Itera sobre las celdas en la columna y convierte los valores de decimal a porcentaje
for celda in columna:
    if isinstance(celda.value, float):  # verifica si el valor de la celda es un decimal
        celda.value = celda.value * 1  # convierte el valor a porcentaje
        celda.number_format = '0%'  # establece el formato de número de la celda como porcentaje con dos decimales
    if (re.search("-", str(celda.value))):
        celda.value = str(int(celda.value * 100)).replace('-','(') + '%)'

#################################################################################
# Repitiendo cabeceras        
# Obtiene el número total de filas y columnas
num_rows = ws.max_row
num_cols = ws.max_column

# Establecer el estilo de fuente y alineación para las cabeceras
font = Font(name='Calibri', size=6, bold=True)
alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

#Cambiar la region en el nombre del archivo dependiendo del archivo de origen
region=''
if re.match("^ReporteOAGLATAM",originFile):
    region = " LATAM"
elif re.match("^ReporteOAGUS",originFile):
    region = " USA"

################################################################################

# Crear un objeto HeaderFooter y asignarle los textos
ws.oddHeader.center.text = "OAG Schedule Competitive Summary" + region + '\n&"Calibri"&8Sunday ' +  formatted_date
ws.oddHeader.center.size = 14
ws.oddHeader.center.font = "Calibri,Bold"

ws.oddFooter.left.text = "Prev Snap:  " + formatted_date3 + " New Snap: " + formatted_date2 + '&R&"Calibri"&8&P - &N'
ws.oddFooter.left.size = 8
ws.oddHeader.left.font = "Calibri"


# Indicar el número de columna que deseas eliminar
num_columna = 1
num_columna2 = 1
num_columna17 = 14
num_columna18 = 15
num_columna19 = 14

# Eliminar la columnas
ws.delete_cols(num_columna)
ws.delete_cols(num_columna2)
ws.delete_cols(num_columna17)
ws.delete_cols(num_columna18)
ws.delete_cols(num_columna19)

# Guardar el archivo con modificación de scale al 110%
ws.page_setup.scale = 110

#Repetir las cabeceras en excel sin agregarlas a los datos
ws.print_title_rows = '1:1'

# Guardar el archivo
wb.save('OAG Schedule Competitive Summary Sunday, ' + formatted_date + region + ' .xlsx')

#Pasar el formato a html
xlsx2html('OAG Schedule Competitive Summary Sunday, ' + formatted_date + region + ' .xlsx', 'OAG Schedule Competitive Summary Sunday, ' + formatted_date + region + ' .html')

#Modificar el codigo html para ajustarlo al formato
#Leer archivo html
with open('OAG Schedule Competitive Summary Sunday, ' + formatted_date + region + ' .html', "r") as f:
    text = f.read()

#Definir el formato de la tabla, cabeceras y pie de pagina
soup = BeautifulSoup(text, "html.parser")
new_div = soup.new_tag("style")
new_div.string = " table {font-family: Calibri; transform: scale(0.85,1); margin-left: auto; margin-right: auto;} @page{size: 8.3in 11.8in;} div.header {display: block; text-align: center; position: running(header); font-family: Calibri; transform: scale(1, 1);} div.footer { display: block; text-align: left; position: running(footer); font-family: Calibri; transform: scale(1, 1); } @page { @top-center { content: element(header) }} @page { @bottom-left { content: element(footer) }} @page {@bottom-right{content: counter(page)' - 'counter(pages); font-family: Calibry; font-size: 10.0px; margin-right: -55px;}}"
soup.html.insert(1, new_div)

#Insertar tag para cabecera
new_divH = soup.new_tag("div")
soup.html.body.insert(1, new_divH)
soup.select('div')[0]['class'] = 'header'

#Insertar tag en cabecera con el titulo del reporte
new_divH1 = soup.new_tag("h1")
new_divH1.string = "OAG Schedule Competitive Summary " + region
soup.html.body.div.insert(1, new_divH1)
soup.select('h1')[0]['style'] = 'font-size: 19.0px; font-weight: normal;'

#Insertar tag en cabecera con fecha del reporte
new_divH2 = soup.new_tag("p")
new_divH2.string = "Sunday " + formatted_date
soup.html.body.div.insert(2, new_divH2)
soup.select('p')[0]['style'] = 'font-size: 13.0px'

#Insertar tag de pie de pagina
new_divF = soup.new_tag("div")
new_divF.string ="Prev Snap:  " + formatted_date3 + " New Snap: " + formatted_date2
soup.html.body.insert(2, new_divF)
soup.select('div')[1]['class'] = 'footer'
soup.select('div')[1]['style'] = 'font-size: 10.0px; margin-left: -55px;'

#Agregar cabeceras de tabla en cada página
style_left = "background-color: #BFBFBF;border-bottom: none;border-collapse: collapse;border-left: none;border-right: none;border-top: none;font-size: 9.0px;font-weight: bold;height: 17.25pt;text-align: left; word-spacing: 20px; line-height: 130%;; transform: scale(1.2,1);"
style_center = "background-color: #BFBFBF;border-bottom: none;border-collapse: collapse;border-left: none;border-right: none;border-top: none;font-size: 9.0px;font-weight: bold;height: 17.25pt;text-align: center; word-spacing: 20px; line-height: 130%;; transform: scale(1.2,1);"

total_rows = num_rows + int(num_rows/79)+2
for max_row in range(80, total_rows, 79):
        original_tag=soup.new_tag("tr")

        new_td1 = soup.new_tag("td", id="Sheet!A0", style=style_left)
        new_td1.string = "CA"
        original_tag.append(new_td1)

        new_td2 = soup.new_tag("td", id="Sheet!B0", style=style_left)
        new_td2.string = "Market"
        original_tag.append(new_td2)

        new_td3 = soup.new_tag("td", id="Sheet!C0", style=style_left)
        new_td3.string = "Ind AM"
        original_tag.append(new_td3)

        new_td4 = soup.new_tag("td", id="Sheet!D0", style=style_left)
        new_td4.string = "Region"
        original_tag.append(new_td4)

        new_td5 = soup.new_tag("td", id="Sheet!E0", style=style_left)
        new_td5.string = "Month"
        original_tag.append(new_td5)

        new_td6 = soup.new_tag("td", id="Sheet!F0", style=style_left)
        new_td6.string = "Chg"
        original_tag.append(new_td6)

        new_td7 = soup.new_tag("td", id="Sheet!G0", style=style_center)
        new_td7.string = "Prev Ops"
        original_tag.append(new_td7)

        new_td8 = soup.new_tag("td", id="Sheet!H0", style=style_center)
        new_td8.string = "New Ops"
        original_tag.append(new_td8)

        new_td9 = soup.new_tag("td", id="Sheet!I0", style=style_center)
        new_td9.string = "Ops Chg"
        original_tag.append(new_td9)

        new_td10 = soup.new_tag("td", id="Sheet!J0", style=style_center)
        new_td10.string = "Prev Seat"
        original_tag.append(new_td10)

        new_td11 = soup.new_tag("td", id="Sheet!K0", style=style_center)
        new_td11.string = "New Seat"
        original_tag.append(new_td11)

        new_td12 = soup.new_tag("td", id="Sheet!L0", style=style_center)
        new_td12.string = "Seat Chg"
        original_tag.append(new_td12)

        new_td13 = soup.new_tag("td", id="Sheet!M0", style=style_center)
        new_td13.string = "%_Seat Chg"
        original_tag.append(new_td13)

        soup.html.body.table.insert(max_row * 2, original_tag)

last_tag = soup.find_all("tr")[-1]
if last_tag.select("td")[0]["id"] == "Sheet!A0":
    last_tag.extract()

#Cambiar el formato de los datos
for link in soup.findAll('td'):
    #Ajustar el tamaño de letra de las celdas
    link['style'] = link['style'].replace('font-size: 6.0px','font-size: 9.0px') 

    #Ajustar el tamaño de letra de las cabeceras 
    link['style'] = link['style'].replace('font-size: 11.0px','font-size: 5.0px')

    #Ajustar el alto de los renglones en blanco (los que van despues de los totales)
    link['style'] = link['style'].replace("background-color: #FFFFFF;border-bottom: none;border-collapse: collapse;border-left-color: #BFBFBF;border-left-style: solid;border-left-width: 1px;border-right-color: #F2F2F2;border-right-style: solid;border-right-width: 1px;border-top: none;font-size: 5.0px;height: 7.5pt","background-color: #FFFFFF;border-bottom: none;border-collapse: collapse;border-left-color: #BFBFBF;border-left-style: solid;border-left-width: 1px;border-right-color: #F2F2F2;border-right-style: solid;border-right-width: 1px;border-top: none;font-size: 5.0px;height: 5.7pt;")    
    link['style'] = link['style'].replace("background-color: #FFFFFF;border-bottom: none;border-collapse: collapse;border-left: none;border-right-color: #F2F2F2;border-right-style: solid;border-right-width: 1px;border-top: none;font-size: 5.0px;height: 7.5pt","background-color: #FFFFFF;border-bottom: none;border-collapse: collapse;border-left: none;border-right-color: #F2F2F2;border-right-style: solid;border-right-width: 1px;border-top: none;font-size: 5.0px;height: 5.7pt")
    link['style'] = link['style'].replace("background-color: #FFFFFF;border-bottom: none;border-collapse: collapse;border-left: none;border-right-color: #BFBFBF;border-right-style: solid;border-right-width: 1px;border-top: none;font-size: 5.0px;height: 7.5pt","background-color: #FFFFFF;border-bottom: none;border-collapse: collapse;border-left: none;border-right-color: #BFBFBF;border-right-style: solid;border-right-width: 1px;border-top: none;font-size: 5.0px;height: 5.7pt")
    
    #Ajustar las cabeceras para que el texto aparezca en dos renglones y ajustar el interlineado
    if "bold;height: 17.25pt;text-align: " in link['style']:
        link['style'] += "; word-spacing: 20px; line-height: 130%;"
    
    #Ajustar la escala de la letra de las celdas
    link['style'] += "; transform: scale(1.2,1);"

    #Agregar padding para que los datos en las celdas no se corten
    if not (re.match('^Sheet!M', link['id'])):
        link['style'] += " padding-right: 9px;"

#Guardar los cambios en el archivo html
with open('OAG Schedule Competitive Summary Sunday, ' + formatted_date + region + ' .html', "w") as f:
    f.write(soup.prettify(formatter="html"))

#Exportar html a pdf
HTML('OAG Schedule Competitive Summary Sunday, ' + formatted_date + region + ' .html').write_pdf('OAG Schedule Competitive Summary Sunday, ' + formatted_date + region + ' .pdf')
